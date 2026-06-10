import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def text_to_number(text):
    """
    Zamienia tekst reprezentujący liczbę na liczbę (int lub float).
    Obsługuje formaty typu '29,364', '1 234,56', '17.199,00', '12345'.
    Zwraca int, jeśli liczba jest całkowita, lub float, jeśli są miejsca po przecinku.
    Jeśli wejście nie jest stringiem, zwraca je bez zmian.
    """
    if not isinstance(text, str):
        return text  # Jeśli to już liczba, zwróć bez zmian
    cleaned = text.replace(' ', '').replace('.', '').replace(',', '.')
    try:
        number = float(cleaned)
        if number.is_integer():
            return int(number)
        else:
            return number
    except ValueError:
        return text



def format_excel_file(input_file, output_file):
    """
    Formatuje plik Excel zgodnie z wymaganiami:
    - Sprawdza komórkę B5 czy zawiera 'C' lub 'B'
    - Formatuje szerokość kolumn różnie dla taryf B i C
    - Formatuje liczby zgodnie z wymaganiami dla każdej taryfy
    - Ustawia odpowiednie style dla komórek
    - Dodaje krawędzie do określonego zakresu
    """
    
    # Wczytaj plik Excel
    wb = openpyxl.load_workbook(input_file)
    
    # Przejdź przez wszystkie arkusze
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Przetwarzam arkusz: {sheet_name}")
        
        # Sprawdź komórkę B5 (wiersz 5, kolumna B)
        b5_value = ws['B5'].value
        if b5_value is None:
            print(f"  Komórka B5 w arkuszu {sheet_name} jest pusta, pomijam.")
            continue
            
        b5_str = str(b5_value)
        print(f"  Wartość B5: {b5_str}")
        
        # Sprawdź czy B5 zawiera 'C' lub 'B'
        if 'C' in b5_str or 'B' in b5_str:
            print(f"  Arkusz {sheet_name} będzie formatowany (B5 zawiera C lub B)")
            
            # Określ typ taryfy
            is_tariff_b = 'B' in b5_str and 'C' not in b5_str
            is_tariff_c = 'C' in b5_str
            
            print(f"  Typ taryfy: {'B' if is_tariff_b else 'C' if is_tariff_c else 'Nieznana'}")
            
            # Ustawianie szerokości kolumn w zależności od taryfy
            if is_tariff_b:
                column_widths = {
                    'A': 12 + 1,
                    'B': 15 + 1,
                    'C': 35 + 1,
                    'D': 14 + 1,
                    'E': 12 + 1,
                    'F': 12 + 1,
                    'G': 14 + 1
                }
            else:  # taryfa C
                column_widths = {
                    'A': 12 + 1,    
                    'B': 20 + 1,
                    'C': 14 + 1,
                    'D': 12 + 1,
                    'E': 14 + 1
                }
            
            # Ustaw szerokość kolumn
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # CZĘŚĆ 1: Przetwarzanie zakresu B10:G21 - usuwanie kropek z liczb
            print("  Usuwam kropki z liczb w zakresie B10:G21...")
            for row in range(10, 22):  # 10-21 włącznie
                for col in range(2, 8):  # B-G (kolumny 2-7)
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.value = text_to_number(cell.value)
                        # Sprawdź czy to liczba z kropkami (np. "17.199,00")
                            # Usuń wszystkie kropki
                        if cell.value is not None and isinstance(cell.value, str):
                            new_value = cell.value.replace('.', '')
                            cell.value = new_value
                            print(f"    {get_column_letter(col)}{row}: {cell.value} -> {new_value}")

            
            # CZĘŚĆ 2: Formatowanie komórek E9 i F9 / B9 i D9
            if is_tariff_b:
                ws['E9'].value = "Zużycie [kWh]"
                ws['F9'].value = "Zużycie [kWh]"
                print("  Zmieniono E9 i F9 na 'Zużycie [kWh]'")
            else:  # taryfa C
                ws['B9'].value = "Zużycie [kWh]"
                ws['D9'].value = "Zużycie [kWh]"
                print("  Zmieniono B9 i D9 na 'Zużycie [kWh]'")
            
            # CZĘŚĆ 3: Przetwarzanie liczb w kolumnach E i F (taryfa B) lub B i D (taryfa C)
            if is_tariff_b:
                # Dla taryfy B: kolumny E i F - usuń przecinki
                for row in range(10, 22):
                    for col in [5, 6]:
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None and isinstance(cell.value, str):
                            if ',' in cell.value:
                                new_value = cell.value.replace(',', '')
                                cell.value = int(new_value) if new_value.isdigit() else new_value
                                print(f"    {col}{row}: usunięto przecinek")
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                           cell.value = cell.value * 1000
                           print(f"    {col}{row}: usunięto przecinek")
            else:
                # Dla taryfy C: kolumny B i D - usuń przecinki
                for row in range(10, 22):
                    for col in [2, 4]:
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None and isinstance(cell.value, str):
                            if ',' in cell.value:
                                new_value = cell.value.replace(',', '')
                                cell.value = int(new_value) if new_value.isdigit() else new_value
                                print(f"    {col}{row}: usunięto przecinek")
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                           cell.value = cell.value * 1000
                           print(f"    {col}{row}: usunięto przecinek")
            
            # CZĘŚĆ 4: Ustawianie formatów komórek dla wierszy 10-21
            for row in range(10, 22):
                for col in range(1, 8):  # A-G
                    cell = ws.cell(row=row, column=col)
                    col_letter = get_column_letter(col)
                    
                    if is_tariff_b:
                        # Formatowanie dla taryfy B
                        if col_letter in ['A', 'B']:
                            # Format ogólny - bez specjalnego formatowania liczb
                            pass
                        elif col_letter == 'C':
                            # Liczbowe z 4 liczbami po przecinku
                            cell.number_format = '#,##0.0000'
                        elif col_letter in ['D', 'G']:
                            # Liczbowe z 2 liczbami po przecinku i separatorem 1000
                            cell.number_format = '#,##0.00'
                        elif col_letter in ['E', 'F']:
                            # Liczbowe z 0 liczbami po przecinku i separatorem 1000
                            cell.number_format = '#,##0'
                    else:
                        # Formatowanie dla taryfy C
                        if col_letter == 'A':
                            # Format ogólny
                            pass
                        elif col_letter in ['C', 'E']:
                            # Liczbowe z 2 liczbami po przecinku i separatorem 1000
                            cell.number_format = '#,##0.00'
                        elif col_letter in ['B', 'D']:
                            # Liczbowe z 0 liczbami po przecinku i separatorem 1000
                            cell.number_format = '#,##0'
            
            # CZĘŚĆ 5: Ustawianie stylów dla wszystkich komórek (oprócz wierszy 1-5)
            for row in ws.iter_rows(min_row=6):
                for cell in row:
                    if cell.value is not None:
                        # Ustaw czcionkę Calibri 11pt, wyśrodkowanie
                        cell.font = Font(name='Calibri', size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # CZĘŚĆ 6: Dodawanie krawędzi do zakresu A8:G21
            # Definiuj styl krawędzi
            thin_border = Side(border_style="thin", color="000000")
            border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
            
            # Zastosuj krawędzie do wszystkich komórek w zakresie A8:G21
            if is_tariff_b:
                for row in range(8, 22):  # 8-21 włącznie
                    for col in range(1, 8):  # A-G + zakres dat 
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
            else:
                for row in range(8, 22):  # 8-21 włącznie
                    for col in range(1, 6):  # A-G
                        cell = ws.cell(row=row, column=col)
                        cell.border = border


            print(f"  Dodano krawędzie do zakresu A8:G21")

            # CZĘŚĆ 7: Pogrubienie tekstu do zakresu A8:G9
            bold_font = Font(bold=True)
            if is_tariff_b:
                for row in range(8, 10):  # 8 i 9
                    for col in range(1, 9):  # A-G to kolumny 1-7
                        cell = ws.cell(row=row, column=col)
                        cell.font = bold_font
            else:
                for row in range(8, 10):  # 8 i 9
                    for col in range(1, 7):  # A-G to kolumny 1-7
                        cell = ws.cell(row=row, column=col)
                        cell.font = bold_font

            if is_tariff_b:
                if ws['H9'].value is not None: 
                    for row in range(9, 22):
                        cell = ws.cell(row=row, column=8)
                        cell.border = border
            else:
                if ws['F9'].value is not None: 
                    for row in range(9, 22):
                        cell = ws.cell(row=row, column=6)
                        cell.border = border
                


            # CZĘŚĆ 8: usuniecie starych linkow do plikow
            row = 24
            max_row = ws.max_row
            while row <= max_row:
                cell_a = ws.cell(row=row, column=1).value
                cell_b = ws.cell(row=row, column=2).value
                if cell_a is None or str(cell_a).strip() == "":
                    break  # Koniec usuwania, bo A jest puste
                # Jeśli w A lub B jest tekst (nie puste)
                if (isinstance(cell_a, str) and cell_a.strip() != "") or (isinstance(cell_b, str) and cell_b.strip() != ""):
                    ws.delete_rows(row, 1)
                    max_row -= 1  # Po usunięciu wiersza max_row się zmniejsza
                    # Nie zwiększamy row, bo po usunięciu kolejny wiersz przesuwa się na miejsce usuniętego
                else:
                    row += 1


        else:
            print(f"  Arkusz {sheet_name} nie będzie formatowany (B5 nie zawiera C ani B)")
    
    # Zapisz sformatowany plik
    wb.save(output_file)
    print(f"\nPlik został zapisany jako: {output_file}")

# Główna funkcja programu
def main():
    input_file = "C:/Users/Jedyn/Downloads/pliki/2025.xlsx"  # Nazwa pliku wejściowego
    output_file = "C:/Users/Jedyn/Downloads/pliki/2025_formatted.xlsx"  # Nazwa pliku wyjściowego
    
    try:
        format_excel_file(input_file, output_file)
        print("\n=== FORMATOWANIE ZAKOŃCZONE POMYŚLNIE ===")
        print(f"Plik źródłowy: {input_file}")
        print(f"Plik wynikowy: {output_file}")
        
    except FileNotFoundError:
        print(f"BŁĄD: Nie znaleziono pliku {input_file}")
        print("Upewnij się, że plik znajduje się w tym samym folderze co skrypt.")
        
    except Exception as e:
        print(f"BŁĄD podczas formatowania: {str(e)}")



if __name__ == "__main__":
    main()
