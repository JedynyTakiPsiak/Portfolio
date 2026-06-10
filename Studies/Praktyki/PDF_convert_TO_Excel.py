import os
import re
import sys
import logging
from datetime import datetime
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Set, Optional, Tuple
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_polish_number(val: str) -> float:
    """Zamienia liczbę w polskim formacie (np. '12.345,67') na float 12345.67"""
    if not val or str(val).strip() == '':
        return 0.0
    val = str(val).replace(' ', '').replace('\xa0', '')  # usuń spacje i twarde spacje
    # najpierw usuń kropki (separatory tysięcy), potem zamień przecinek na kropkę (dziesiętne)
    val = val.replace('.', '').replace(',', '.')
    try:
        return float(val)
    except Exception:
        return 0.0

def format_number_for_excel(val, places=4) -> str:
    """Zamienia float na polski string z przecinkiem dziesiętnym"""
    try:
        num = float(val)
        # Zaokrąglij do 4 miejsc po przecinku
        formatted = f"{num:.4f}".rstrip('0').rstrip('.')
        return formatted.replace('.', ',')
    except Exception:
        return str(val)

@dataclass
class Measurement:
    """Struktura danych dla pojedynczego pomiaru"""
    year: int
    month: int
    period: str
    tariff: str
    dist_max_power: str = ''
    dist_reactive_energy: str = ''
    dist_usage: str = ''
    dist_net: str = ''
    sales_usage: str = ''
    sales_net: str = ''
    source_file: str = ''
    is_correction: bool = False

@dataclass
class PPEEntry:
    """Struktura danych dla punktu poboru energii"""
    ppe_number: str
    address: str = ''
    contract_power: str = ''
    meter_number: str = ''
    tariff: str = ''
    measurements: List[Measurement] = field(default_factory=list)
    source_files: Set[str] = field(default_factory=set)

class PPEAggregator:
    """Klasa do agregacji danych PPE"""
    def __init__(self, logger):
        self.entries: Dict[str, PPEEntry] = {}
        self.pdf_folder_name: str = ""
        self.logger = logger

    def set_pdf_folder(self, folder_path: str):
        """Ustawia nazwę folderu PDF dla hiperłączy"""
        self.pdf_folder_name = os.path.basename(folder_path)

    def add_measurement(self, ppe_number: str, address: str, contract_power: str, 
                       meter_number: str, tariff: str, measurement: Measurement):
        """Dodaje pomiar do PPE z obsługą priorytetyzacji korekty"""
        entry = self.entries.get(ppe_number)
        if not entry:
            entry = PPEEntry(ppe_number, address, contract_power, meter_number, tariff)
            self.entries[ppe_number] = entry
        else:
            # Aktualizuj dane bazowe jeśli są lepsze
            if address and not entry.address:
                entry.address = address
            if contract_power and not entry.contract_power:
                entry.contract_power = contract_power
            if meter_number and not entry.meter_number:
                entry.meter_number = meter_number
            if tariff and not entry.tariff:
                entry.tariff = tariff

        # Obsługa priorytetyzacji korekty
        existing_measurement = None
        for i, existing in enumerate(entry.measurements):
            if existing.year == measurement.year and existing.month == measurement.month:
                existing_measurement = existing
                existing_index = i
                break
        
        if existing_measurement:
            # Jeśli nowy pomiar to korekta, zastąp istniejący
            if measurement.is_correction:
                self.logger.info(f"Zastępuję pomiar korektą dla PPE {ppe_number}, {measurement.year}/{measurement.month}")
                entry.measurements[existing_index] = measurement
            # Jeśli istniejący to korekta, a nowy nie - zostaw istniejący
            elif existing_measurement.is_correction:
                self.logger.info(f"Pomijam pomiar (już mam korektę) dla PPE {ppe_number}, {measurement.year}/{measurement.month}")
                pass
            else:
                # Oba nie są korektą - zastąp (ostatni wygrywa)
                entry.measurements[existing_index] = measurement
        else:
            # Brak istniejącego pomiaru - dodaj nowy
            entry.measurements.append(measurement)

        entry.source_files.add(measurement.source_file)

    def get_years(self) -> Set[int]:
        """Zwraca lata występujące w danych"""
        years = set()
        for entry in self.entries.values():
            for measurement in entry.measurements:
                years.add(measurement.year)
        return years

    def all_entries(self) -> List[PPEEntry]:
        """Zwraca wszystkie wpisy PPE"""
        return list(self.entries.values())

class PDFProcessor:
    """Klasa do przetwarzania plików PDF"""
    
    def __init__(self, logger):
        self.logger = logger
        self.month_names = {
            'styczeń': 1, 'styczen': 1, 'stycznia': 1,
            'luty': 2, 'lutego': 2,
            'marzec': 3, 'marca': 3,
            'kwiecień': 4, 'kwiecien': 4, 'kwietnia': 4,
            'maj': 5, 'maja': 5,
            'czerwiec': 6, 'czerwca': 6,
            'lipiec': 7, 'lipca': 7,
            'sierpień': 8, 'sierpien': 8, 'sierpnia': 8,
            'wrzesień': 9, 'wrzesien': 9, 'września': 9,
            'październik': 10, 'pazdziernik': 10, 'października': 10,
            'listopad': 11, 'listopada': 11,
            'grudzień': 12, 'grudzien': 12, 'grudnia': 12
        }

    def extract_text_from_pdf(self, pdf_path: str) -> str:
        """Wyciąga tekst z pliku PDF"""
        text = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            self.logger.error(f"Błąd wyciągania tekstu z {pdf_path}: {e}")
            return ""
        return text

    def detect_invoice_type(self, text: str) -> str:
        """Rozpoznaje typ faktury"""
        if "ZA USŁUGI DYSTRYBUCJI" in text:
            return "DYSTRYBUCJA"
        elif "ZA ENERGIĘ ELEKTRYCZNĄ I USŁUGI DYSTRYBUCJI" in text:
            return "PELNA"
        elif "ZA ENERGIĘ ELEKTRYCZNĄ" in text:
            return "SPRZEDAZ"
        else:
            return "NIEZNANY"

    def split_invoices(self, text: str) -> List[str]:
        """Dzieli tekst na pojedyncze faktury wg 'FAKTURA VAT'"""
        parts = text.split("FAKTURA VAT")
        invoices = []
        
        if parts[0].strip():
            invoices.append(parts[0])
            
        for part in parts[1:]:
            if part.strip():
                invoices.append("FAKTURA VAT" + part)
                
        return invoices

    def check_if_corrected(self, text: str) -> bool:
        """Sprawdza czy to faktura korygująca"""
        return ("KOREKTA" in text or "FAKTURA VAT KOREKTA" in text) and "Należało policzyć" in text

    def extract_corrected_section(self, text: str) -> str:
        """LOGIKA Z test16.py: Wyciąga TYLKO sekcję korekty"""
        if "Należało policzyć" not in text:
            return text
            
        # Podziel tekst na sekcje
        sections = text.split("Należało policzyć")
        if len(sections) < 2:
            return text
            
        # Weź tylko część "Należało policzyć" i wszystko po niej
        corrected_section = "Należało policzyć" + sections[1]
        
        self.logger.info("UŻYWAM WYŁĄCZNIE SEKCJI 'Należało policzyć' - IGNORUJĘ 'Policzono'")
        return corrected_section

    def extract_ppe_sections(self, text: str) -> List[str]:
        """Dzieli tekst na sekcje PPE"""
        sections = []
        lines = text.split('\n')
        current_section = []
        
        for line in lines:
            if re.match(r'Miejsce poboru energii\s*\d*\.?', line) and current_section:
                sections.append('\n'.join(current_section))
                current_section = [line]
            else:
                current_section.append(line)
        
        if current_section:
            sections.append('\n'.join(current_section))
        
        return sections

    def extract_period(self, section: str) -> Optional[Tuple[datetime, datetime, str]]:
        """Wyciąga okres rozliczeniowy"""
        patterns = [
            r'Za okres od\s*(\d{1,2})[./](\d{1,2})[./](\d{4})\s*do\s*(\d{1,2})[./](\d{1,2})[./](\d{4})',
            r'od\s*(\d{1,2})[./](\d{1,2})[./](\d{4})\s*do\s*(\d{1,2})[./](\d{1,2})[./](\d{4})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, section)
            if match:
                try:
                    d1 = datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
                    d2 = datetime(int(match.group(6)), int(match.group(5)), int(match.group(4)))
                    period_str = f"{d1.strftime('%d/%m/%Y')} – {d2.strftime('%d/%m/%Y')}"
                    return d1, d2, period_str
                except ValueError:
                    continue
        
        return None

    def clean_address(self, raw_address: str) -> str:
        """Czyści adres z numeracji i zbędnych elementów"""
        if not raw_address:
            return ""
        
        cleaned = re.sub(r'^\d+\.\s*,?\s*', '', raw_address.strip())
        cleaned = re.sub(r'\s+Za okres.*$', '', cleaned)
        cleaned = re.sub(r'^[,\s]+', '', cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        return cleaned.strip()

    def extract_basic_info(self, section: str) -> Dict[str, str]:
        """Wyciąga podstawowe informacje o PPE"""
        info = {
            'ppe_number': '',
            'address': '',
            'tariff': '',
            'contract_power': '',
            'meter_number': ''
        }
        
        # Kod PPE
        ppe_match = re.search(r'Kod PPE:\s*(\d{14,18})', section)
        if ppe_match:
            info['ppe_number'] = ppe_match.group(1)
        
        # Adres
        address_patterns = [
            r'Miejsce poboru energii\s*\d*\.?\s*\n([^\n]+?)(?:\s+Za okres|$)',
            r'Miejsce poboru energii\s*\d*\.?\s*([^,\n]+(?:,[^,\n]+)*?)(?:\s+Za okres|\n|$)',
            r'Miejsce poboru energii\s*(\d+\.\s*[^,\n]+(?:,[^,\n]+)*?)(?:\s+Za okres|\n|$)'
        ]
        
        for pattern in address_patterns:
            match = re.search(pattern, section, re.MULTILINE)
            if match:
                raw_address = match.group(1).strip()
                info['address'] = self.clean_address(raw_address)
                break
        
        # Taryfa
        tariff_patterns = [
            r'Grupa taryfowa:\s*([BC]\d{1,2})',
            r'taryfowa:\s*([BC]\d{1,2})',
            r'Taryfa:\s*([BC]\d{1,2})',
            r'([BC]\d{1,2})\s*-?\s*(?:grupa|taryfa)'
        ]
        
        for pattern in tariff_patterns:
            match = re.search(pattern, section, re.I)
            if match:
                info['tariff'] = match.group(1).upper()
                break
        
        # Moc umowna
        power_patterns = [
            r'Moc umowna:\s*([0-9,\.]+)\s*kW',
            r'Moc umowna:\s*([0-9,\.]+)',
            r'umowna:\s*([0-9,\.]+)'
        ]
        
        for pattern in power_patterns:
            match = re.search(pattern, section)
            if match:
                info['contract_power'] = format_number_for_excel(parse_polish_number(match.group(1)))
                break
        
        # Numer licznika
        meter_patterns = [
            r'Wskaźnik rozliczeniowy\s+mocy\s+nr\s+([A-Z0-9\-\/]+)',
            r'Wskaźnik rozliczeniowy.*?nr\s+([A-Z0-9\-\/]+)',
            r'Licznik rozliczeniowy.*?nr\s+([A-Z0-9\-\/]+)',
            r'Nr licznika:\s*([A-Z0-9\-\/]+)',
            r'licznika\s*([A-Z0-9]{8,})',
            r'nr\s+([A-Z0-9]{8,})'
        ]
        
        for pattern in meter_patterns:
            match = re.search(pattern, section, re.I)
            if match:
                meter_num = match.group(1).strip()
                if len(meter_num) >= 5:
                    info['meter_number'] = meter_num
                    break
        
        return info

    def extract_table_section(self, text: str, table_name: str) -> str:
        """Wyciąga sekcję tabeli o danej nazwie"""
        patterns = [
            rf'{re.escape(table_name)}.*?(?=(?:ROZLICZENIE|ODCZYTY|Razem do zapłaty|PODSUMOWANIE|Strona \d+|$))',
            rf'{table_name}.*?(?=\n[A-ZĄĆĘŁŃÓŚŹŻ]{{2,}}|\n\d{{2}}\.\d{{2}}\.\d{{4}}|$)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.DOTALL | re.I)
            if match:
                return match.group(0)
        
        return ""

    def extract_max_power(self, section: str) -> str:
        """LOGIKA Z test16.py: Wyciąga maksymalną moc z kolumny Wskazanie bieżące"""
        odczyty_section = self.extract_table_section(section, "ODCZYTY")
        if not odczyty_section:
            self.logger.debug("Nie znaleziono sekcji ODCZYTY")
            return ""
        
        # Znajdź sekcję wskaźnika rozliczeniowego mocy
        wskaznik_match = re.search(
            r'Wskaźnik rozliczeniowy\s+mocy.*?(?=Licznik rozliczeniowy|Wskaźnik rozliczeniowy energii|ROZLICZENIE|$)', 
            odczyty_section, 
            re.DOTALL | re.I
        )
        
        if not wskaznik_match:
            self.logger.debug("Nie znaleziono sekcji wskaźnika rozliczeniowego mocy")
            return ""
        
        wskaznik_section = wskaznik_match.group(0)
        
        # Analizuj linie zawierające "Moc pobrana" i wyciągnij wartości z kolumny "Wskazanie bieżące"
        power_values = []
        lines = wskaznik_section.split('\n')
        
        for line in lines:
            if 'Moc pobrana' in line:
                self.logger.debug(f"Analizuję linię mocy: {line}")
                
                # Podziel linię na kolumny po białych znakach
                columns = re.split(r'\s+', line.strip())
                self.logger.debug(f"Kolumny w linii: {columns}")
                
                # Wskazanie bieżące to zazwyczaj 3 kolumna (indeks 2)
                if len(columns) >= 4:
                    for col_idx in [2, 3, 4]:
                        if col_idx < len(columns):
                            try:
                                value_str = columns[col_idx]
                                if re.match(r'^[\d,\.]+$', value_str):
                                    value = parse_polish_number(value_str)
                                    if 1 <= value <= 10000:
                                        power_values.append(value)
                                        self.logger.debug(f"Dodano wartość mocy: {value} kW z kolumny {col_idx}")
                                        break
                            except (ValueError, IndexError):
                                continue
        
        if power_values:
            max_power = max(power_values)
            self.logger.info(f"Maksymalna moc pobrana: {max_power} kW")
            return format_number_for_excel(str(max_power))
        
        self.logger.debug("Nie znaleziono prawidłowych wartości mocy")
        return ""

    def extract_reactive_energy(self, section: str) -> str:
        """LOGIKA Z test16.py: Wyciąga energię bierną pojemnościową"""
        odczyty_section = self.extract_table_section(section, "ODCZYTY")
        if not odczyty_section:
            return ""
        
        # Znajdź sekcję energii biernej pojemnościowej
        reactive_match = re.search(
            r'Licznik rozliczeniowy\s+energii biernej pojemnościowej.*?(?=Rozliczenie|ROZLICZENIE|Wskaźnik rozliczeniowy|$)',
            odczyty_section, 
            re.DOTALL | re.I
        )
        
        if not reactive_match:
            return ""
        
        reactive_section = reactive_match.group(0)
        lines = reactive_section.split('\n')
        
        # Sprawdź najpierw całodobową
        for line in lines:
            if 'całodobowa' in line.lower():
                self.logger.debug(f"Analizuję linię energii biernej całodobowej: {line}")
                
                # Podziel linię na kolumny
                columns = re.split(r'\s+', line.strip())
                self.logger.debug(f"Kolumny w linii: {columns}")
                
                # Wskazanie bieżące to zazwyczaj 3 kolumna (indeks 2)
                if len(columns) >= 4:
                    for col_idx in [2, 3, 4]:
                        if col_idx < len(columns):
                            try:
                                value_str = columns[col_idx]
                                if re.match(r'^[\d,\.]+$', value_str):
                                    value = parse_polish_number(value_str)
                                    if 0.1 <= value <= 100000:
                                        self.logger.info(f"Energia bierna całodobowa: {value} kWh")
                                        return format_number_for_excel(str(value))
                            except (ValueError, IndexError):
                                continue
        
        # Jeśli nie znaleziono całodobowej, sumuj z trzech stref
        zone_patterns = ['szczyt przedpołudniowy', 'szczyt popołudniowy', 'pozostałe godziny doby']
        total_value = 0.0
        values_found = 0
        
        for zone_pattern in zone_patterns:
            for line in lines:
                if zone_pattern in line.lower():
                    self.logger.debug(f"Analizuję strefę: {zone_pattern} - {line}")
                    columns = re.split(r'\s+', line.strip())
                    if len(columns) >= 4:
                        for col_idx in [2, 3, 4]:
                            if col_idx < len(columns):
                                try:
                                    value_str = columns[col_idx]
                                    if re.match(r'^[\d,\.]+$', value_str):
                                        value = parse_polish_number(value_str)
                                        if 0.1 <= value <= 100000:
                                            self.logger.debug(f"Dodaję do sumy: {value} kWh ze strefy {zone_pattern}")
                                            total_value += value
                                            values_found += 1
                                            break
                                except (ValueError, IndexError):
                                    continue
                    break
        
        if values_found > 0:
            self.logger.info(f"Energia bierna suma z {values_found} stref: {total_value} kWh")
            return format_number_for_excel(str(total_value))
        
        return ""

    def extract_distribution_usage_from_summary(self, section: str) -> str:
        """Wyciąga zużycie z podsumowania dla faktur dystrybucyjnych"""
        usage_patterns = [
            r'Zużycie:\s*([0-9,.]+)\s*kWh',
            r'Zużycie:\s*([0-9,.]+)\s*MWh'
        ]
        
        for pattern in usage_patterns:
            match = re.search(pattern, section, re.I)
            if match:
                value_str = match.group(1)
                value = parse_polish_number(value_str)
                
                if 'MWh' in pattern:
                    self.logger.info(f"Zużycie dystrybucji z podsumowania: {value} MWh")
                    return format_number_for_excel(str(value))
                else:  # kWh
                    value_mwh = value / 1000
                    self.logger.info(f"Zużycie dystrybucji z podsumowania: {value} kWh = {value_mwh} MWh")
                    return format_number_for_excel(str(value_mwh))
        
        return ""

    def extract_distribution_net_from_summary(self, section: str) -> str:
        """Wyciąga wartość netto z podsumowania dla faktur dystrybucyjnych"""
        net_patterns = [
            r'Ogółem wartość netto:\s*([0-9,.]+)\s*zł',
            r'Ogółem wartość netto:\s*([0-9,.]+)',
            r'wartość netto:\s*([0-9,.]+)'
        ]
        
        for pattern in net_patterns:
            match = re.search(pattern, section, re.I)
            if match:
                value_str = match.group(1)
                value = parse_polish_number(value_str)
                self.logger.info(f"Wartość netto dystrybucji z podsumowania: {value} PLN")
                return format_number_for_excel(str(value))
        
        return ""

    def extract_distribution_usage_from_balancing(self, section: str) -> str:
        """LOGIKA Z test16.py: Wyciąga zużycie dystrybucji z 'Zużycie po bilansowaniu' lub 'Zużycie'"""
        dist_section = self.extract_table_section(section, "ROZLICZENIE - USŁUGA DYSTRYBUCJI ENERGII")
        if not dist_section:
            return ""
        
        # PRIORYTET 1: Szukaj "Zużycie po bilansowaniu: X kWh"
        balancing_patterns = [
            r'Zużycie po bilansowaniu:\s*([0-9,.]+)\s*kWh',
            r'Zużycie po bilansowaniu:\s*([0-9,.]+)\s*MWh',
            r'po bilansowaniu:\s*([0-9,.]+)\s*kWh',
            r'po bilansowaniu:\s*([0-9,.]+)\s*MWh'
        ]
        
        for pattern in balancing_patterns:
            match = re.search(pattern, dist_section, re.I)
            if match:
                value_str = match.group(1)
                value = parse_polish_number(value_str)
                
                if 'MWh' in pattern:
                    self.logger.info(f"Zużycie dystrybucji po bilansowaniu: {value} MWh")
                    return format_number_for_excel(str(value))
                else:  # kWh
                    value_mwh = value / 1000
                    self.logger.info(f"Zużycie dystrybucji po bilansowaniu: {value} kWh = {value_mwh} MWh")
                    return format_number_for_excel(str(value_mwh))
        
        # PRIORYTET 2: Szukaj "Zużycie: X kWh" jeśli nie ma bilansowania
        regular_patterns = [
            r'Zużycie:\s*([0-9,.]+)\s*kWh',
            r'Zużycie:\s*([0-9,.]+)\s*MWh'
        ]
        
        for pattern in regular_patterns:
            match = re.search(pattern, dist_section, re.I)
            if match:
                value_str = match.group(1)
                value = parse_polish_number(value_str)
                
                if 'MWh' in pattern:
                    self.logger.info(f"Zużycie dystrybucji: {value} MWh")
                    return format_number_for_excel(str(value))
                else:  # kWh
                    value_mwh = value / 1000
                    self.logger.info(f"Zużycie dystrybucji: {value} kWh = {value_mwh} MWh")
                    return format_number_for_excel(str(value_mwh))
        
        return ""

    def extract_distribution_usage(self, section: str, invoice_type: str) -> str:
        """LOGIKA Z test16.py: Wyciąga zużycie energii z dystrybucji"""
        # Dla faktur dystrybucyjnych używaj podsumowania
        if invoice_type == "DYSTRYBUCJA":
            return self.extract_distribution_usage_from_summary(section)
        
        # Dla pełnych faktur używaj "Zużycie po bilansowaniu" lub "Zużycie"
        elif invoice_type == "PELNA":
            return self.extract_distribution_usage_from_balancing(section)
        
        # Dla innych typów używaj standardowej metody
        dist_section = self.extract_table_section(section, "ROZLICZENIE - USŁUGA DYSTRYBUCJI ENERGII")
        if not dist_section:
            return ""
        
        # Sprawdź najpierw czy jest "Zużycie po bilansowaniu" lub "Zużycie"
        usage_from_balancing = self.extract_distribution_usage_from_balancing(section)
        if usage_from_balancing:
            return usage_from_balancing
        
        # Fallback: szukaj w sekcji opłaty zmiennej sieciowej
        usage_patterns = [
            r'Opłata zmienna sieciowa\s+całodobowa\s+kWh\s+[^0-9]*([0-9,.]+)',
            r'całodobowa\s+kWh\s+[^0-9]*([0-9,.]+)',
            r'zmienna sieciowa.*?całodobowa.*?kWh.*?([0-9,.]+)',
        ]
        
        for pattern in usage_patterns:
            match = re.search(pattern, dist_section, re.I | re.M)
            if match:
                numbers = re.findall(r'([0-9]+[,.][0-9]+|[0-9]+)', match.group(0))
                if numbers:
                    for num_str in numbers:
                        value_kwh = parse_polish_number(num_str)
                        if 1 <= value_kwh <= 1000000:
                            value_mwh = value_kwh / 1000
                            self.logger.info(f"Zużycie dystrybucji z opłaty zmiennej: {value_kwh} kWh = {value_mwh} MWh")
                            return format_number_for_excel(str(value_mwh))
        
        return ""

    def extract_distribution_net(self, section: str, invoice_type: str) -> str:
        """LOGIKA Z test16.py: Wyciąga wartość netto dystrybucji"""
        # Dla faktur dystrybucyjnych używaj podsumowania
        if invoice_type == "DYSTRYBUCJA":
            return self.extract_distribution_net_from_summary(section)
        
        # Dla wszystkich innych typów używaj standardowej metody
        dist_section = self.extract_table_section(section, "ROZLICZENIE - USŁUGA DYSTRYBUCJI ENERGII")
        if not dist_section:
            return ""
        
        # Szukaj "Ogółem wartość - usługa dystrybucji: X"
        net_patterns = [
            r'Ogółem wartość - usługa dystrybucji:\s*([0-9,.]+)',
            r'usługa dystrybucji:\s*([0-9,.]+)',
            r'wartość - usługa dystrybucji.*?([0-9,.]+)'
        ]
        
        for pattern in net_patterns:
            match = re.search(pattern, dist_section, re.I)
            if match:
                value_str = match.group(1)
                value = parse_polish_number(value_str)
                self.logger.info(f"Wartość netto dystrybucji: {value} PLN")
                return format_number_for_excel(str(value))
        
        return ""

    def extract_sales_usage(self, section: str) -> str:
        """Wyciąga zużycie energii ze sprzedaży"""
        sales_section = self.extract_table_section(section, "ROZLICZENIE - SPRZEDAŻ ENERGII")
        if not sales_section:
            return ""
        
        # Szukaj energii elektrycznej czynnej
        usage_patterns = [
            r'Energia elektryczna czynna\s+całodobowa\s+MWh\s+([0-9,.]+)',
            r'Energia elektryczna czynna\s+całodobowa\s+kWh\s+([0-9,.]+)',
            r'całodobowa\s+MWh\s+([0-9,.]+)',
            r'całodobowa\s+kWh\s+([0-9,.]+)',
        ]
        
        for pattern in usage_patterns:
            match = re.search(pattern, sales_section, re.I | re.M)
            if match:
                numbers = re.findall(r'([0-9]+[,.][0-9]+|[0-9]+)', match.group(0))
                if numbers:
                    for num_str in numbers:
                        value = parse_polish_number(num_str)
                        if 0.001 <= value <= 1000000:
                            if 'MWh' in pattern:
                                self.logger.info(f"Zużycie sprzedaży: {value} MWh")
                                return format_number_for_excel(str(value))
                            else:  # kWh
                                value_mwh = value / 1000
                                self.logger.info(f"Zużycie sprzedaży: {value} kWh = {value_mwh} MWh")
                                return format_number_for_excel(str(value_mwh))
        
        return ""

    def extract_sales_net(self, section: str) -> str:
        """Wyciąga wartość netto sprzedaży"""
        sales_section = self.extract_table_section(section, "ROZLICZENIE - SPRZEDAŻ ENERGII")
        if not sales_section:
            return ""
        
        net_patterns = [
            r'Ogółem wartość - sprzedaż energii:\s*([0-9,.]+)',
            r'sprzedaż energii:\s*([0-9,.]+)',
            r'wartość - sprzedaż.*?([0-9,.]+)'
        ]
        
        for pattern in net_patterns:
            match = re.search(pattern, sales_section, re.I)
            if match:
                value_str = match.group(1)
                value = parse_polish_number(value_str)
                self.logger.info(f"Wartość netto sprzedaży: {value} PLN")
                return format_number_for_excel(str(value))
        
        return ""

    def process_pdf(self, pdf_path: str, aggregator: PPEAggregator, 
                   file_index: int, total_files: int) -> int:
        """LOGIKA Z test16.py: Przetwarza pojedynczy plik PDF"""
        print(f"\rPrzetwarzam plik {os.path.basename(pdf_path)} {file_index}/{total_files}", 
              end='', flush=True)
        
        text = self.extract_text_from_pdf(pdf_path)
        if not text:
            return 0
        
        processed_count = 0
        invoices = self.split_invoices(text)
        
        for invoice_text in invoices:
            # Rozpoznanie typu faktury
            invoice_type = self.detect_invoice_type(invoice_text)
            is_corrected = self.check_if_corrected(invoice_text)
            
            self.logger.info(f"Typ faktury: {invoice_type}, Korekta: {is_corrected}")
            
            # Dla korekt używaj WYŁĄCZNIE sekcji "Należało policzyć"
            if is_corrected:
                invoice_text = self.extract_corrected_section(invoice_text)
            
            sections = self.extract_ppe_sections(invoice_text)
            
            for section in sections:
                if not section.strip():
                    continue
                
                basic_info = self.extract_basic_info(section)
                if not basic_info['ppe_number']:
                    continue
                
                period_info = self.extract_period(section)
                if not period_info:
                    continue
                
                start_date, end_date, period_str = period_info
                year = end_date.year
                month = end_date.month
                
                tariff = basic_info['tariff']
                
                measurement = Measurement(
                    year=year,
                    month=month,
                    period=period_str,
                    tariff=tariff,
                    source_file=os.path.basename(pdf_path),
                    is_correction=is_corrected
                )
                
                # Różne dane w zależności od typu faktury
                if invoice_type == "DYSTRYBUCJA":
                    # Faktura za usługi dystrybucji - tylko dane dystrybucji
                    if tariff.startswith('B'):
                        measurement.dist_max_power = self.extract_max_power(section)
                        measurement.dist_reactive_energy = self.extract_reactive_energy(section)
                    measurement.dist_usage = self.extract_distribution_usage(section, invoice_type)
                    measurement.dist_net = self.extract_distribution_net(section, invoice_type)
                    measurement.sales_usage = ''
                    measurement.sales_net = ''
                    
                elif invoice_type == "SPRZEDAZ":
                    # Faktura za energię elektryczną - tylko dane sprzedaży
                    measurement.dist_max_power = ''
                    measurement.dist_reactive_energy = ''
                    measurement.dist_usage = ''
                    measurement.dist_net = ''
                    measurement.sales_usage = self.extract_sales_usage(section)
                    measurement.sales_net = self.extract_sales_net(section)
                    
                elif invoice_type == "PELNA":
                    # NOWA LOGIKA PELNA: Pełna faktura - zużycie dystrybucji = zużycie sprzedaży
                    if tariff.startswith('B'):
                        measurement.dist_max_power = self.extract_max_power(section)
                        measurement.dist_reactive_energy = self.extract_reactive_energy(section)
                    elif tariff.startswith('C'):
                        measurement.dist_max_power = ''
                        measurement.dist_reactive_energy = ''
                    
                    # Dane dystrybucji
                    measurement.dist_usage = self.extract_distribution_usage(section, invoice_type)
                    measurement.dist_net = self.extract_distribution_net(section, invoice_type)
                    
                    # NOWA LOGIKA: Jeśli jest zużycie dystrybucji, kopiuj je jako zużycie sprzedaży
                    if measurement.dist_usage:
                        measurement.sales_usage = measurement.dist_usage
                        self.logger.info(f"PELNA: Kopiuję zużycie dystrybucji ({measurement.dist_usage}) jako zużycie sprzedaży")
                    else:
                        measurement.sales_usage = self.extract_sales_usage(section)
                    
                    # Wartość netto sprzedaży - standardowo
                    measurement.sales_net = self.extract_sales_net(section)
                    
                else:  # NIEZNANY
                    # Standardowa logika
                    if tariff.startswith('B'):
                        measurement.dist_max_power = self.extract_max_power(section)
                        measurement.dist_reactive_energy = self.extract_reactive_energy(section)
                    elif tariff.startswith('C'):
                        measurement.dist_max_power = ''
                        measurement.dist_reactive_energy = ''
                    
                    measurement.dist_usage = self.extract_distribution_usage(section, invoice_type)
                    measurement.dist_net = self.extract_distribution_net(section, invoice_type)
                    measurement.sales_usage = self.extract_sales_usage(section)
                    measurement.sales_net = self.extract_sales_net(section)
                
                # Loguj szczegółowe informacje
                self.logger.info(f"=== POMIAR PPE {basic_info['ppe_number']} ===")
                self.logger.info(f"Typ faktury: {invoice_type}")
                self.logger.info(f"Adres: {basic_info['address']}")
                self.logger.info(f"Nr licznika: {basic_info['meter_number']}")
                self.logger.info(f"Taryfa: {tariff}")
                self.logger.info(f"Moc umowna: {basic_info['contract_power']}")
                self.logger.info(f"Okres: {period_str}")
                self.logger.info(f"Rok/Miesiąc: {year}/{month}")
                if tariff.startswith('B'):
                    self.logger.info(f"Moc max: {measurement.dist_max_power}")
                    self.logger.info(f"Energia bierna: {measurement.dist_reactive_energy}")
                self.logger.info(f"Zużycie dystr: {measurement.dist_usage}")
                self.logger.info(f"Netto dystr: {measurement.dist_net}")
                self.logger.info(f"Zużycie sprzedaż: {measurement.sales_usage}")
                self.logger.info(f"Netto sprzedaż: {measurement.sales_net}")
                self.logger.info(f"Plik źródłowy: {pdf_path}")
                self.logger.info(f"Typ: {'KOREKTA' if is_corrected else 'NORMALNA'}")
                self.logger.info("=" * 50)
                
                aggregator.add_measurement(
                    basic_info['ppe_number'],
                    basic_info['address'],
                    basic_info['contract_power'],
                    basic_info['meter_number'],
                    tariff,
                    measurement
                )
                
                processed_count += 1
        
        return processed_count

class ExcelExporter:
    """UKŁAD Z test15.py: Klasa do eksportu danych do plików Excel"""
    
    def __init__(self, base_filename: str, pdf_folder_name: str, logger):
        self.base_filename = base_filename
        self.pdf_folder_name = pdf_folder_name
        self.workbooks: Dict[int, Workbook] = {}
        self.logger = logger

    def get_workbook(self, year: int) -> Workbook:
        """Zwraca lub tworzy workbook dla danego roku"""
        if year not in self.workbooks:
            self.workbooks[year] = Workbook()
            if 'Sheet' in self.workbooks[year].sheetnames:
                self.workbooks[year].remove(self.workbooks[year]['Sheet'])
        return self.workbooks[year]

    def apply_formatting(self, ws, start_row: int, end_row: int, num_cols: int):
        """Stosuje formatowanie do arkusza"""
        # Kolory i style
        header_font = Font(bold=True, size=11, color="000000")
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        data_font = Font(size=10)
        
        # Obramowanie
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatuj nagłówki
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=9, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Formatuj nagłówki grup (wiersz 8)
        for col in range(2, num_cols + 1):
            cell = ws.cell(row=8, column=col)
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        # Formatuj dane
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        # Ustaw szerokości kolumn
        column_widths = [15, 18, 25, 18, 18, 18, 18]
        for i, width in enumerate(column_widths[:num_cols], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def export_ppe_to_year(self, ppe: PPEEntry, year: int):
        """UKŁAD Z test15.py: Eksportuje PPE do arkusza w odpowiednim roku"""
        wb = self.get_workbook(year)
        sheet_name = f'PPE_{ppe.ppe_number}'
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name)
        
        # Nagłówki informacyjne
        ws['A1'] = 'PPE:'
        ws['B1'] = ppe.ppe_number
        ws['A2'] = 'Adres:'
        ws['B2'] = ppe.address
        ws['A3'] = 'Nr licznika:'
        ws['B3'] = ppe.meter_number
        ws['A4'] = 'Moc umowna:'
        ws['B4'] = ppe.contract_power
        ws['A5'] = 'Taryfa:'
        ws['B5'] = ppe.tariff
        
        # Różne układy dla taryfy B i C (z test15.py)
        is_tariff_b = ppe.tariff and ppe.tariff.startswith('B')
        
        if is_tariff_b:
            # Układ dla taryfy B (pełna tabela)
            ws['B8'] = 'Świadczenie usługi dystrybucji'
            ws['F8'] = 'Sprzedaż energii'
            ws.merge_cells('B8:E8')
            ws.merge_cells('F8:G8')
            
            headers = [
                'Miesiąc', 
                'Moc max [kW]', 
                'Energia bierna (pojemnościowa) [kWh]',
                'Cena netto [PLN]', 
                'Zużycie [MWh]', 
                'Zużycie [MWh]', 
                'Cena netto [PLN]'
            ]
        else:
            # Układ dla taryfy C (bez mocy max i energii biernej)
            ws['B8'] = 'Świadczenie usługi dystrybucji'
            ws['D8'] = 'Sprzedaż energii'
            ws.merge_cells('B8:C8')
            ws.merge_cells('D8:E8')
            
            headers = [
                'Miesiąc',
                'Zużycie [MWh]',
                'Cena netto [PLN]', 
                'Zużycie [MWh]', 
                'Cena netto [PLN]'
            ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=9, column=i, value=header)
        
        # Przygotuj dane dla każdego miesiąca (1-12)
        month_data = {}
        for measurement in ppe.measurements:
            if measurement.year == year:
                month_data[measurement.month] = measurement
        
        # Wypełnij wiersze dla miesięcy 1-12
        for month in range(1, 13):
            row = 9 + month
            ws.cell(row=row, column=1, value=month)
            
            if month in month_data:
                measurement = month_data[month]
                
                if is_tariff_b:
                    # Taryfa B: pełny układ
                    ws.cell(row=row, column=2, value=measurement.dist_max_power)
                    ws.cell(row=row, column=3, value=measurement.dist_reactive_energy)
                    ws.cell(row=row, column=4, value=measurement.dist_net)
                    ws.cell(row=row, column=5, value=measurement.dist_usage)
                    ws.cell(row=row, column=6, value=measurement.sales_usage)
                    ws.cell(row=row, column=7, value=measurement.sales_net)
                else:
                    # Taryfa C: pomijamy kolumny mocy max i energii biernej
                    ws.cell(row=row, column=2, value=measurement.dist_usage)
                    ws.cell(row=row, column=3, value=measurement.dist_net)
                    ws.cell(row=row, column=4, value=measurement.sales_usage)
                    ws.cell(row=row, column=5, value=measurement.sales_net)
        
        # Lista plików źródłowych
        ws.cell(row=24, column=1, value='Lista plików źródłowych:')
        ws.cell(row=24, column=1).font = Font(bold=True)
        
        ws.cell(row=25, column=1, value='Lp.')
        ws.cell(row=25, column=2, value='Nazwa pliku')
        ws.cell(row=25, column=1).font = Font(bold=True)
        ws.cell(row=25, column=2).font = Font(bold=True)
        
        # Filtruj pliki dla danego roku
        year_files = set()
        for measurement in ppe.measurements:
            if measurement.year == year:
                year_files.add(measurement.source_file)
        
        for i, file_name in enumerate(sorted(year_files), 1):
            row = 25 + i
            ws.cell(row=row, column=1, value=i)
            cell = ws.cell(row=row, column=2, value=file_name)
            cell.hyperlink = f'{self.pdf_folder_name}/{file_name}'
            cell.font = Font(color="0000FF", underline="single")
        
        # Zastosuj formatowanie
        num_cols = 7 if is_tariff_b else 5
        self.apply_formatting(ws, 10, 21, num_cols)

    def save_all(self):
        """Zapisuje wszystkie pliki Excel"""
        for year, wb in self.workbooks.items():
            filename = f"{self.base_filename}_{year}.xlsx"
            wb.save(filename)
            print(f"\nZapisano: {filename}")

def main():
    """Główna funkcja programu"""
    print("=== AUTOMATYZACJA PRZETWARZANIA FAKTUR ENERGETYCZNYCH ===")
    
    pdf_folder = input("Podaj ścieżkę do folderu z plikami PDF: ").strip()
    excel_file = input("Podaj ścieżkę do pliku Excel (wynik): ").strip()
    
    if not os.path.exists(pdf_folder):
        print(f"BŁĄD: Folder {pdf_folder} nie istnieje!")
        return
    
    # Przygotuj nazwę bazową pliku (bez rozszerzenia)
    base_filename = os.path.splitext(excel_file)[0]
    
    # Konfiguracja logowania z nazwą log_file
    log_file = f"{base_filename}_log.txt"
    log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    logger = logging.getLogger("PPE")
    logger.setLevel(logging.INFO)
    
    # Usuń istniejące handlery
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(log_formatter)
    logger.addHandler(file_handler)
    
    aggregator = PPEAggregator(logger)
    aggregator.set_pdf_folder(pdf_folder)
    processor = PDFProcessor(logger)
    
    pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print("BŁĄD: Nie znaleziono plików PDF w podanym folderze!")
        return
    
    print(f"Znaleziono {len(pdf_files)} plików PDF")
    
    total_processed = 0
    error_count = 0
    
    for idx, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(pdf_folder, pdf_file)
        
        try:
            processed_count = processor.process_pdf(pdf_path, aggregator, idx, len(pdf_files))
            total_processed += processed_count
        except Exception as e:
            logger.error(f"BŁĄD w pliku {pdf_file}: {e}")
            error_count += 1
            continue
    
    print(f"\n\nPrzetworzono {total_processed} pomiarów z {len(pdf_files)} plików")
    print(f"Błędy: {error_count} plików")
    print(f"Utworzono {len(aggregator.entries)} unikalnych PPE")
    
    years = aggregator.get_years()
    print(f"Znalezione lata: {sorted(years)}")
    
    exporter = ExcelExporter(base_filename, aggregator.pdf_folder_name, logger)
    
    exported_count = 0
    for ppe in aggregator.all_entries():
        ppe_years = set(m.year for m in ppe.measurements)
        for year in ppe_years:
            try:
                exporter.export_ppe_to_year(ppe, year)
                exported_count += 1
            except Exception as e:
                logger.error(f"Błąd eksportu PPE {ppe.ppe_number} dla roku {year}: {e}")
    
    exporter.save_all()
    
    print(f"ZAKOŃCZONO!")
    print(f"Wyeksportowano {exported_count} arkuszy PPE")
    print(f"Log szczegółowy: {log_file}")

if __name__ == "__main__":
    main()
